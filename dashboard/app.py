"""
SAP Migration Post-Load Validator - Live Dashboard
Run:  python dashboard/app.py
Open: http://localhost:5000
"""

import sys
import threading
import time
import json
from pathlib import Path
from datetime import datetime

from flask import Flask, render_template, jsonify, send_file, request, Response
from werkzeug.utils import secure_filename

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.validator import MaterialValidator
from core.reporter  import generate_excel_report
from core.field_labels import get_label, load_custom_labels, SAP_FIELD_LABELS


app = Flask(__name__)

BASE_DIR    = Path(__file__).parent.parent
REPORTS_DIR = BASE_DIR / "reports"
CONFIG_FILE = BASE_DIR / "config.json"
LABELS_FILE = BASE_DIR / "custom_labels.csv"

REPORTS_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_CONFIG = {
    "source_dir":      str(BASE_DIR / "data" / "source"),
    "target_dir":      str(BASE_DIR / "data" / "target"),
    "pass_threshold":  100.0,
    "selected_fields": [],
    "manual_pairs":    [],
}

results_store = {}

scan_status = {
    "last_scan":       None,
    "scanning":        False,
    "error":           None,
    "current_file":    None,
    "total_files":     0,
    "completed_files": 0,
}

file_states  = {}
activity_log = []

SUPPORTED_EXT = {".csv", ".xlsx", ".xls"}
scan_lock     = threading.Lock()


# ── Config helpers ────────────────────────────────────────────────────────────

def load_config():
    if CONFIG_FILE.exists():
        try:
            return {**DEFAULT_CONFIG, **json.loads(CONFIG_FILE.read_text())}
        except Exception as e:
            print(f"Config load failed: {e}")
    return dict(DEFAULT_CONFIG)


def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2))


def get_dirs():
    cfg = load_config()
    src = Path(cfg.get("source_dir", DEFAULT_CONFIG["source_dir"]))
    tgt = Path(cfg.get("target_dir", DEFAULT_CONFIG["target_dir"]))
    src.mkdir(parents=True, exist_ok=True)
    tgt.mkdir(parents=True, exist_ok=True)
    return src, tgt


def log_event(message, level="info"):
    entry = {
        "ts":      datetime.now().strftime("%H:%M:%S"),
        "message": message,
        "level":   level,
    }
    activity_log.append(entry)
    if len(activity_log) > 50:
        activity_log.pop(0)
    print(f"  [{entry['ts']}] {message}")


def cleanup_old_reports(keep_latest=20):
    files = sorted(
        REPORTS_DIR.glob("*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    for old_file in files[keep_latest:]:
        try:
            old_file.unlink()
        except Exception as e:
            log_event(f"Could not delete old report {old_file.name}: {e}", "warn")


def _get_custom_labels():
    """Load custom labels from file if it exists, else empty dict."""
    return load_custom_labels(str(LABELS_FILE)) if LABELS_FILE.exists() else {}


def resolve_label(field_name: str) -> str:
    """
    Resolve a SAP technical field name to its English description.
    e.g. KUNNR -> Customer Number
         MATNR -> Material Number
         KTOKD -> Customer Account Group
    Uses custom labels first, then built-in SAP dictionary.
    """
    custom = _get_custom_labels()
    return get_label(field_name, custom)


# ── File discovery ────────────────────────────────────────────────────────────

def get_available_files():
    SOURCE_DIR, TARGET_DIR = get_dirs()
    src_files = sorted(
        [f for f in SOURCE_DIR.iterdir() if f.suffix.lower() in SUPPORTED_EXT],
        key=lambda f: f.name.upper(),
    )
    tgt_files = sorted(
        [f for f in TARGET_DIR.iterdir() if f.suffix.lower() in SUPPORTED_EXT],
        key=lambda f: f.name.upper(),
    )
    return src_files, tgt_files


def discover_pairs():
    SOURCE_DIR, TARGET_DIR = get_dirs()
    cfg          = load_config()
    manual_pairs = cfg.get("manual_pairs", [])

    src_files = {
        f.name: f
        for f in SOURCE_DIR.iterdir()
        if f.suffix.lower() in SUPPORTED_EXT
    }
    tgt_files = {
        f.name: f
        for f in TARGET_DIR.iterdir()
        if f.suffix.lower() in SUPPORTED_EXT
    }

    pairs    = []
    used_src = set()
    used_tgt = set()

    for mp in manual_pairs:
        src_name = mp.get("source_file", "")
        tgt_name = mp.get("target_file", "")
        name     = mp.get("name", "").upper().strip() or Path(src_name).stem.upper()
        sp = str(src_files[src_name]) if src_name in src_files else None
        tp = str(tgt_files[tgt_name]) if tgt_name in tgt_files else None
        has_pair = sp is not None and tp is not None
        mtime    = (
            max(Path(sp).stat().st_mtime, Path(tp).stat().st_mtime)
            if has_pair else None
        )
        pairs.append({
            "name": name, "source_path": sp, "target_path": tp,
            "has_pair": has_pair, "mtime": mtime,
            "source_file": Path(sp).name if sp else src_name,
            "target_file": Path(tp).name if tp else tgt_name,
            "match_type": "manual",
            "missing": [] if has_pair else (
                (["source"] if not sp else []) + (["target"] if not tp else [])
            ),
        })
        if sp: used_src.add(src_name)
        if tp: used_tgt.add(tgt_name)

    src_by_stem = {}
    for fname, fpath in src_files.items():
        if fname not in used_src:
            src_by_stem[Path(fname).stem.upper()] = (fname, fpath)

    tgt_by_stem = {}
    for fname, fpath in tgt_files.items():
        if fname not in used_tgt:
            tgt_by_stem[Path(fname).stem.upper()] = (fname, fpath)

    for stem in sorted(set(src_by_stem) & set(tgt_by_stem)):
        src_fname, src_fpath = src_by_stem[stem]
        tgt_fname, tgt_fpath = tgt_by_stem[stem]
        mtime = max(src_fpath.stat().st_mtime, tgt_fpath.stat().st_mtime)
        pairs.append({
            "name": stem, "source_path": str(src_fpath), "target_path": str(tgt_fpath),
            "has_pair": True, "mtime": mtime,
            "source_file": src_fname, "target_file": tgt_fname,
            "match_type": "auto", "missing": [],
        })
        used_src.add(src_fname)
        used_tgt.add(tgt_fname)

    for fname, fpath in src_files.items():
        if fname not in used_src:
            pairs.append({
                "name": Path(fname).stem.upper(), "source_path": str(fpath),
                "target_path": None, "has_pair": False, "mtime": None,
                "source_file": fname, "target_file": None,
                "match_type": "unmatched", "missing": ["target"],
            })

    for fname, fpath in tgt_files.items():
        if fname not in used_tgt:
            pairs.append({
                "name": Path(fname).stem.upper(), "source_path": None,
                "target_path": str(fpath), "has_pair": False, "mtime": None,
                "source_file": None, "target_file": fname,
                "match_type": "unmatched", "missing": ["source"],
            })

    return pairs


# ── Business status ───────────────────────────────────────────────────────────

def calculate_business_status(result, pass_threshold):
    ss          = result.summary_stats
    pass_rate   = float(ss.get("pass_rate_pct", 0))
    only_source = int(result.records_only_in_source or 0)
    only_target = int(result.records_only_in_target or 0)

    if pass_rate < pass_threshold:
        return {
            "status": "FAIL", "field_status": "FAIL", "record_status": "CHECKED",
            "message": (
                f"Field validation failed. Pass rate is {pass_rate:.2f}% "
                f"which is below threshold {pass_threshold:.2f}%."
            ),
        }

    if only_source > 0 or only_target > 0:
        if only_source > 0 and only_target > 0:
            record_msg = (
                f"{only_source:,} records exist only in source and "
                f"{only_target:,} records exist only in target."
            )
        elif only_target > 0:
            record_msg = f"Target has {only_target:,} extra records not found in source."
        else:
            record_msg = f"Source has {only_source:,} records not found in target."
        return {
            "status": "WARNING", "field_status": "PASS", "record_status": "WARNING",
            "message": (
                f"Field validation passed with {pass_rate:.2f}% "
                f"against threshold {pass_threshold:.2f}%, but {record_msg}"
            ),
        }

    return {
        "status": "PASS", "field_status": "PASS", "record_status": "PASS",
        "message": (
            f"Validation passed. Field pass rate is {pass_rate:.2f}% "
            f"and source/target records are fully reconciled."
        ),
    }


# ── Core: get fields for a table (used by dashboard field selector) ────────────

def get_fields_for_object(name: str, source_path: str = None, target_path: str = None) -> list:
    """
    Return a labelled field list for a given object name.
    Priority:
      1. Actual columns from files (if paths provided)
      2. Object config key_fields
      3. All SAP known fields for that object type
    Each entry: {field, label, in_source, in_target, common, selected}
    """
    from core.object_config import get_object_config
    cfg        = load_config()
    sel_set    = set(cfg.get("selected_fields", []))
    obj_cfg    = get_object_config(name)
    key_fields = obj_cfg.get("key_fields", [])

    custom = _get_custom_labels()

    def make_entry(col, in_src, in_tgt):
        return {
            "field":     col,
            "label":     get_label(col, custom),   # KUNNR -> Customer Number etc
            "in_source": in_src,
            "in_target": in_tgt,
            "common":    in_src and in_tgt,
            "selected":  len(sel_set) == 0 or col in sel_set,
        }

    # If we have actual files, read their real columns
    if source_path and target_path:
        try:
            src_cols, tgt_cols = _read_file_headers(source_path, tgt_path=target_path)
            src_set = set(src_cols)
            tgt_set = set(tgt_cols)
            common   = sorted(src_set & tgt_set)
            src_only = sorted(src_set - tgt_set)
            tgt_only = sorted(tgt_set - src_set)
            fields = []
            for c in common:   fields.append(make_entry(c, True,  True))
            for c in src_only: fields.append(make_entry(c, True,  False))
            for c in tgt_only: fields.append(make_entry(c, False, True))
            return fields
        except Exception:
            pass

    # Fall back to object config key fields
    if key_fields:
        return [make_entry(f, True, True) for f in key_fields]

    # Last resort: return all known SAP fields
    return [make_entry(f, True, True) for f in SAP_FIELD_LABELS]


def _read_file_headers(src_path: str, tgt_path: str = None):
    """Read just the header row from CSV/XLSX — very fast, no full load."""
    import csv

    def headers(path, delim=","):
        p = Path(path)
        if not p.exists():
            return []
        if p.suffix.lower() in (".xlsx", ".xls"):
            import openpyxl
            wb   = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            ws   = wb.active
            cols = [str(c.value).strip().upper()
                    for c in next(ws.iter_rows(max_row=1)) if c.value]
            wb.close()
            return cols
        with open(str(p), encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=delim)
            return [c.strip().upper() for c in next(reader)]

    src_cols = headers(src_path) if src_path else []
    tgt_cols = headers(tgt_path) if tgt_path else []
    return src_cols, tgt_cols


# ── Validation runner ─────────────────────────────────────────────────────────

def run_validation(name, source_path, target_path):
    cfg             = load_config()
    pass_threshold  = float(cfg.get("pass_threshold", 100.0))
    selected_fields = cfg.get("selected_fields", [])
    custom          = _get_custom_labels()

    # Auto-detect SAP object config
    from core.object_config import get_object_config
    obj_cfg  = get_object_config(name)
    join_key = obj_cfg.get("join_key", None)

    # Smart default: if no manual selection, use object key fields
    effective_fields = selected_fields
    if not effective_fields and obj_cfg.get("key_fields"):
        effective_fields = obj_cfg["key_fields"]
        log_event(
            f"{name}: auto-selected {len(effective_fields)} key fields "
            f"from {obj_cfg.get('description', name)} config",
            "info",
        )

    # Warn for large files
    src_mb = Path(source_path).stat().st_size / (1024 * 1024)
    tgt_mb = Path(target_path).stat().st_size / (1024 * 1024)
    if src_mb > 50 or tgt_mb > 50:
        log_event(
            f"{name}: large files ({src_mb:.1f} MB / {tgt_mb:.1f} MB) "
            f"— this may take a few minutes",
            "warn",
        )

    validator = MaterialValidator(
        pass_threshold=pass_threshold,
        selected_fields=effective_fields if effective_fields else None,
        join_key=join_key,
        custom_labels=custom if custom else None,
    )

    result          = validator.validate(source_path, target_path)
    ss              = result.summary_stats
    business_status = calculate_business_status(result, pass_threshold)

    # ── Build field rows with full mismatch detail ────────────────────────────
    field_rows = []
    for fr in result.field_results:
        # Resolve English label: KUNNR -> "Customer Number"
        label = get_label(fr.field_source, custom)

        field_rows.append({
            "field":          fr.field_source,
            "field_label":    label,                  # human-readable name
            "field_target":   fr.field_target,
            "type":           "numeric" if fr.is_numeric else "string",
            "tolerance":      fr.tolerance_used,
            "total":          fr.total_records,
            "matched":        fr.matched,
            "mismatched":     fr.mismatched,
            "miss_source":    fr.missing_in_source,
            "miss_target":    fr.missing_in_target,
            "match_pct":      fr.match_pct,
            "pass_threshold": fr.pass_threshold,
            "status":         fr.status,
            # Full mismatch detail — NOT capped here, capped only for dashboard display
            "mismatches":     fr.mismatch_details,
            "mismatch_count": len(fr.mismatch_details),
        })

    # ── Build mapping info with labels ────────────────────────────────────────
    mapping = None
    if result.mapping:
        mapping = {
            "join_key":           result.mapping.join_key,
            "join_key_label":     get_label(result.mapping.join_key, custom),
            "matched_fields":     result.mapping.matched_fields,
            # Every field gets its English label
            "matched_labels": {
                f: get_label(f, custom)
                for f in result.mapping.matched_fields
            },
            "source_only_fields": result.mapping.source_only_fields,
            "source_only_labels": {
                f: get_label(f, custom)
                for f in result.mapping.source_only_fields
            },
            "target_only_fields": result.mapping.target_only_fields,
            "target_only_labels": {
                f: get_label(f, custom)
                for f in result.mapping.target_only_fields
            },
            "numeric_fields":  result.mapping.numeric_fields,
            "tolerance_map":   result.mapping.tolerance_map,
            "selected_fields": result.mapping.selected_fields,
            "pass_threshold":  result.mapping.pass_threshold,
        }

    ts             = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{name}_{ts}.xlsx"
    excel_path     = REPORTS_DIR / excel_filename

    # ── Build available_fields for Settings field selector ────────────────────
    # This is what populates the checkboxes in Settings so the user can
    # pick which fields to validate — with English labels next to SAP codes
    try:
        src_cols, tgt_cols = _read_file_headers(source_path, target_path)
        src_set = set(src_cols)
        tgt_set = set(tgt_cols)
        common   = sorted(src_set & tgt_set)
        src_only = sorted(src_set - tgt_set)
        tgt_only = sorted(tgt_set - src_set)
        sel_set  = set(effective_fields)
        available_fields = []
        for f in common:
            available_fields.append({
                "field":     f,
                "label":     get_label(f, custom),
                "in_source": True, "in_target": True, "common": True,
                "selected":  len(sel_set) == 0 or f in sel_set,
            })
        for f in src_only:
            available_fields.append({
                "field":     f,
                "label":     get_label(f, custom),
                "in_source": True, "in_target": False, "common": False,
                "selected":  False,
            })
        for f in tgt_only:
            available_fields.append({
                "field":     f,
                "label":     get_label(f, custom),
                "in_source": False, "in_target": True, "common": False,
                "selected":  False,
            })
    except Exception:
        available_fields = [
            {"field": fr["field"], "label": fr["field_label"],
             "in_source": True, "in_target": True, "common": True,
             "selected": True}
            for fr in field_rows
        ]

    result_dict = {
        "name":                   name,
        "sap_object":             obj_cfg.get("description", name),
        "status":                 business_status["status"],
        "validator_status":       result.overall_status,
        "field_status":           business_status["field_status"],
        "record_status":          business_status["record_status"],
        "business_message":       business_status["message"],
        "source_file":            Path(source_path).name,
        "target_file":            Path(target_path).name,
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "pass_threshold":         pass_threshold,
        "selected_fields":        effective_fields,
        "errors":                 result.errors,
        "mapping":                mapping,
        "field_results":          field_rows,
        "available_fields":       available_fields,   # for Settings field selector
        "run_at":                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "excel_file":             excel_filename,
    }

    try:
        generate_excel_report(result_dict, str(excel_path))
        cleanup_old_reports()
    except Exception as e:
        result_dict["excel_error"] = str(e)
        log_event(f"Excel failed for {name}: {e}", "error")

    return result_dict


# ── Scan orchestrator ─────────────────────────────────────────────────────────

def scan_and_validate_all():
    if not scan_lock.acquire(blocking=False):
        log_event("Scan already running — skipping duplicate scan", "warn")
        return

    scan_status["scanning"]        = True
    scan_status["error"]           = None
    scan_status["current_file"]    = None
    scan_status["total_files"]     = 0
    scan_status["completed_files"] = 0

    try:
        pairs       = discover_pairs()
        valid_pairs = [p for p in pairs if p["has_pair"]]
        scan_status["total_files"] = len(valid_pairs)

        for pair in pairs:
            name = pair["name"]

            if not pair["has_pair"]:
                prev = file_states.get(name, {})
                if prev.get("state") != "unmatched":
                    side  = "source" if pair["source_path"] else "target"
                    other = "target" if side == "source" else "source"
                    log_event(
                        f"{name}: found in {side} only — waiting for {other} file",
                        "warn",
                    )
                    file_states[name] = {
                        "state":       "unmatched",
                        "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "source_file": pair["source_file"],
                        "target_file": pair["target_file"],
                    }
                continue

            last_mtime = pair["mtime"]
            existing   = results_store.get(name)
            prev_state = file_states.get(name, {})

            if not existing:
                match_type = pair.get("match_type", "auto")
                match_note = f" [{match_type}]" if match_type != "auto" else ""
                log_event(
                    f"{name}: new file pair{match_note} — "
                    f"{pair['source_file']} + {pair['target_file']}",
                    "info",
                )
            elif prev_state.get("_mtime") != last_mtime:
                log_event(f"{name}: file changed — re-validating", "info")
            else:
                scan_status["completed_files"] += 1
                continue

            scan_status["current_file"] = name
            file_states[name] = {
                "state":       "validating",
                "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_file": pair["source_file"],
                "target_file": pair["target_file"],
                "_mtime":      last_mtime,
            }

            try:
                result = run_validation(name, pair["source_path"], pair["target_path"])
                result["_mtime"] = last_mtime
                results_store[name] = result

                file_states[name] = {
                    "state":         "done",
                    "detected_at":   file_states[name]["detected_at"],
                    "validated_at":  result["run_at"],
                    "source_file":   pair["source_file"],
                    "target_file":   pair["target_file"],
                    "_mtime":        last_mtime,
                    "status":        result["status"],
                    "field_status":  result["field_status"],
                    "record_status": result["record_status"],
                    "message":       result["business_message"],
                }

                level = (
                    "success" if result["status"] == "PASS"    else
                    "warn"    if result["status"] == "WARNING"  else
                    "error"
                )
                log_event(
                    f"{name}: {result['status']} — {result['business_message']} "
                    f"| Matched: {result['records_matched']:,} "
                    f"| Source only: {result['records_only_in_source']:,} "
                    f"| Target only: {result['records_only_in_target']:,}",
                    level,
                )

            except Exception as e:
                file_states[name]["state"] = "error"
                file_states[name]["error"] = str(e)
                scan_status["error"]       = str(e)
                log_event(f"{name}: error — {e}", "error")

            finally:
                scan_status["completed_files"] += 1

        scan_status["last_scan"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    except Exception as e:
        scan_status["error"] = str(e)
        log_event(f"Scan error: {e}", "error")

    finally:
        scan_status["scanning"]     = False
        scan_status["current_file"] = None
        scan_lock.release()


def background_watcher(interval=60):
    while True:
        scan_and_validate_all()
        time.sleep(interval)


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("dashboard.html")


@app.route("/api/scan", methods=["POST"])
def api_scan():
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/status")
def api_status():
    pairs           = discover_pairs()
    cfg             = load_config()
    source_dir, target_dir = get_dirs()
    selected_fields = cfg.get("selected_fields", [])

    return jsonify({
        "last_scan":       scan_status["last_scan"],
        "scanning":        scan_status["scanning"],
        "error":           scan_status["error"],
        "current_file":    scan_status["current_file"],
        "total_files":     scan_status["total_files"],
        "completed_files": scan_status["completed_files"],
        "source_dir":      str(source_dir),
        "target_dir":      str(target_dir),
        "pairs":           pairs,
        "file_states":     file_states,
        "total_tables":    len([p for p in pairs if p["has_pair"]]),
        "unmatched":       len([p for p in pairs if not p["has_pair"]]),
        "pass_threshold":  cfg.get("pass_threshold", 100.0),
        "selected_fields": selected_fields,
        "validation_mode": "all_fields" if not selected_fields else "selected_fields",
    })


@app.route("/api/results")
def api_results():
    return jsonify(list(results_store.values()))


@app.route("/api/results/<name>")
def api_result_detail(name):
    r = results_store.get(name.upper())
    if r:
        return jsonify(r)
    return jsonify({"error": "Not found"}), 404


@app.route("/api/activity")
def api_activity():
    return jsonify(list(reversed(activity_log)))


# ── Upload ────────────────────────────────────────────────────────────────────

@app.route("/api/upload/source", methods=["POST"])
def upload_source():
    return _handle_upload(request, get_dirs()[0], "source")


@app.route("/api/upload/target", methods=["POST"])
def upload_target():
    return _handle_upload(request, get_dirs()[1], "target")


def _handle_upload(req, dest_dir, side):
    if "file" not in req.files:
        return jsonify({"error": "No file"}), 400
    saved = []
    for f in req.files.getlist("file"):
        if not f.filename:
            continue
        save_name = secure_filename(f.filename)
        suffix    = Path(save_name).suffix.lower()
        if suffix not in SUPPORTED_EXT:
            return jsonify({"error": f"Unsupported file type: {save_name}"}), 400
        f.save(str(dest_dir / save_name))
        log_event(f"Uploaded to {side}: {save_name}", "info")
        saved.append(save_name)
    if saved:
        threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True, "saved": saved})


@app.route("/api/upload/labels", methods=["POST"])
def upload_labels():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400
    safe_name = secure_filename(f.filename)
    f.save(str(LABELS_FILE))
    log_event(f"Custom labels uploaded: {safe_name}", "info")
    results_store.clear()
    for n in file_states:
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True, "filename": safe_name})


# ── Config ────────────────────────────────────────────────────────────────────

@app.route("/api/config", methods=["GET"])
def api_get_config():
    cfg = load_config()

    # Build available_fields from the most recently validated result
    # so Settings field selector always has labels like "Customer Number (KUNNR)"
    available = []
    if results_store:
        first = next(iter(results_store.values()))
        # Use available_fields from result if present (set during validation)
        if first.get("available_fields"):
            available = first["available_fields"]
        else:
            # Fallback: build from field_results with labels
            custom = _get_custom_labels()
            available = [
                {
                    "field":     fr["field"],
                    "label":     get_label(fr["field"], custom),
                    "in_source": True, "in_target": True, "common": True,
                    "selected":  True,
                }
                for fr in first.get("field_results", [])
            ]

    return jsonify({
        "source_dir":       cfg.get("source_dir",      DEFAULT_CONFIG["source_dir"]),
        "target_dir":       cfg.get("target_dir",      DEFAULT_CONFIG["target_dir"]),
        "pass_threshold":   cfg.get("pass_threshold",  100.0),
        "selected_fields":  cfg.get("selected_fields", []),
        "available_fields": available,
        "labels_file_exists": LABELS_FILE.exists(),
        "labels_file":      str(LABELS_FILE) if LABELS_FILE.exists() else None,
    })


@app.route("/api/config", methods=["POST"])
def api_set_config():
    data    = request.get_json(force=True)
    cfg     = load_config()
    changed = False

    for key in ("source_dir", "target_dir"):
        if key in data and str(data[key]).strip():
            new_path = str(Path(str(data[key]).strip()))
            if new_path != cfg.get(key):
                cfg[key] = new_path
                changed  = True

    if "pass_threshold" in data:
        thr = float(data["pass_threshold"])
        if thr != cfg.get("pass_threshold"):
            cfg["pass_threshold"] = thr
            changed = True
            log_event(f"Pass threshold updated to {thr}%", "info")

    if "selected_fields" in data:
        sel = [str(f).strip().upper() for f in data["selected_fields"] if str(f).strip()]
        if sel != cfg.get("selected_fields", []):
            cfg["selected_fields"] = sel
            changed = True
            log_event(
                f"Field selection updated: {len(sel)} fields" if sel
                else "Field selection: all fields", "info",
            )

    if changed:
        save_config(cfg)
        results_store.clear()
        for n in file_states:
            if file_states[n].get("state") == "done":
                file_states[n]["state"] = "changed"
        threading.Thread(target=scan_and_validate_all, daemon=True).start()

    return jsonify({"ok": True, "config": cfg})


# ── Field preview (Settings → Load fields from files) ─────────────────────────

@app.route("/api/fields/preview", methods=["POST"])
def api_fields_preview():
    data      = request.get_json(force=True)
    src_path  = data.get("source_path", "").strip()
    tgt_path  = data.get("target_path", "").strip()
    src_delim = data.get("source_delimiter", ",")
    tgt_delim = data.get("target_delimiter", ",")

    custom = _get_custom_labels()

    def read_headers(path, delim):
        p = Path(path)
        if not p.exists():
            return None, f"File not found: {path}"
        try:
            if p.suffix.lower() in (".xlsx", ".xls"):
                import openpyxl
                wb   = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                ws   = wb.active
                cols = [str(c.value).strip().upper()
                        for c in next(ws.iter_rows(max_row=1)) if c.value]
                wb.close()
            else:
                import csv
                with open(str(p), encoding="utf-8-sig") as f:
                    reader = csv.reader(f, delimiter=delim)
                    cols   = [c.strip().upper() for c in next(reader)]
            return cols, None
        except Exception as e:
            return None, str(e)

    src_cols, src_err = read_headers(src_path, src_delim) if src_path else ([], None)
    tgt_cols, tgt_err = read_headers(tgt_path, tgt_delim) if tgt_path else ([], None)

    errors = {}
    if src_err: errors["source"] = src_err
    if tgt_err: errors["target"] = tgt_err

    src_set  = set(src_cols or [])
    tgt_set  = set(tgt_cols or [])
    common   = sorted(src_set & tgt_set)
    src_only = sorted(src_set - tgt_set)
    tgt_only = sorted(tgt_set - src_set)

    cfg          = load_config()
    selected_set = set(cfg.get("selected_fields", []))

    fields = []
    for col in common:
        fields.append({
            "field":     col,
            "label":     get_label(col, custom),   # KUNNR -> Customer Number
            "in_source": True, "in_target": True, "common": True,
            "selected":  len(selected_set) == 0 or col in selected_set,
        })
    for col in src_only:
        fields.append({
            "field":     col,
            "label":     get_label(col, custom),
            "in_source": True, "in_target": False, "common": False, "selected": False,
        })
    for col in tgt_only:
        fields.append({
            "field":     col,
            "label":     get_label(col, custom),
            "in_source": False, "in_target": True, "common": False, "selected": False,
        })

    return jsonify({
        "fields":    fields,
        "src_count": len(src_cols or []),
        "tgt_count": len(tgt_cols or []),
        "common":    len(common),
        "src_only":  len(src_only),
        "tgt_only":  len(tgt_only),
        "errors":    errors,
    })


# ── SAP field label lookup (used by dashboard for any field) ──────────────────

@app.route("/api/field-label/<field_name>")
def api_field_label(field_name):
    """Return the English label for a single SAP field code."""
    custom = _get_custom_labels()
    label  = get_label(field_name.upper(), custom)
    return jsonify({
        "field": field_name.upper(),
        "label": label,
        "is_known": label != field_name.upper(),
    })


@app.route("/api/field-labels", methods=["POST"])
def api_field_labels_bulk():
    """
    Bulk label lookup.
    Body: {"fields": ["KUNNR", "MATNR", "KTOKD", ...]}
    Returns: {"KUNNR": "Customer Number", "MATNR": "Material Number", ...}
    """
    data   = request.get_json(force=True)
    fields = [str(f).strip().upper() for f in data.get("fields", [])]
    custom = _get_custom_labels()
    result = {f: get_label(f, custom) for f in fields}
    return jsonify(result)


# ── Manual pair management ────────────────────────────────────────────────────

@app.route("/api/files/list")
def api_files_list():
    src_files, tgt_files = get_available_files()
    return jsonify({
        "source_files": [f.name for f in src_files],
        "target_files": [f.name for f in tgt_files],
    })


@app.route("/api/pairs", methods=["GET"])
def api_pairs_get():
    cfg = load_config()
    return jsonify(cfg.get("manual_pairs", []))


@app.route("/api/pairs", methods=["POST"])
def api_pairs_save():
    data       = request.get_json(force=True)
    pairs      = data.get("pairs", [])
    seen_names = set()
    clean      = []
    for p in pairs:
        name     = str(p.get("name",        "")).strip().upper()
        src_file = str(p.get("source_file", "")).strip()
        tgt_file = str(p.get("target_file", "")).strip()
        if not name or not src_file or not tgt_file:
            continue
        if name in seen_names:
            continue
        seen_names.add(name)
        clean.append({"name": name, "source_file": src_file, "target_file": tgt_file})

    cfg = load_config()
    cfg["manual_pairs"] = clean
    save_config(cfg)
    results_store.clear()
    for n in list(file_states.keys()):
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    log_event(f"Manual pairs updated: {len(clean)} pair(s) saved", "info")
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True, "saved": len(clean)})


@app.route("/api/pairs/<name>", methods=["DELETE"])
def api_pairs_delete(name):
    cfg    = load_config()
    pairs  = cfg.get("manual_pairs", [])
    before = len(pairs)
    pairs  = [p for p in pairs if p["name"].upper() != name.upper()]
    cfg["manual_pairs"] = pairs
    save_config(cfg)
    removed = before - len(pairs)
    if removed:
        results_store.pop(name.upper(), None)
        file_states.pop(name.upper(), None)
        log_event(f"Manual pair removed: {name}", "info")
    return jsonify({"ok": True, "removed": removed})


# ── SAP object catalogue ──────────────────────────────────────────────────────

@app.route("/api/objects")
def api_objects():
    from core.object_config import SAP_OBJECT_CONFIG
    custom = _get_custom_labels()
    return jsonify([
        {
            "key":         k,
            "description": v.get("description", k),
            "join_key":    v.get("join_key", ""),
            "join_key_label": get_label(v.get("join_key", ""), custom),
            # Include English labels for every key field
            "key_fields": [
                {"field": f, "label": get_label(f, custom)}
                for f in v.get("key_fields", [])
            ],
        }
        for k, v in SAP_OBJECT_CONFIG.items()
    ])


# ── Labels / reports / download ───────────────────────────────────────────────

@app.route("/api/labels/sample")
def api_labels_sample():
    lines = ["FIELD_NAME,FRIENDLY_LABEL"]
    for k, v in list(SAP_FIELD_LABELS.items())[:20]:
        lines.append(f"{k},{v}")
    return Response(
        "\n".join(lines).encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=sample_labels.csv"},
    )


@app.route("/api/download/<name>")
def api_download(name):
    r = results_store.get(name.upper())
    if not r:
        return jsonify({"error": "Not found"}), 404
    path = REPORTS_DIR / r.get("excel_file", "")
    if not path.exists():
        return jsonify({"error": "Missing"}), 404
    return send_file(
        str(path), as_attachment=True, download_name=path.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download-file/<filename>")
def api_download_file(filename):
    path = REPORTS_DIR / filename
    if not path.exists() or not filename.endswith(".xlsx"):
        return jsonify({"error": "Not found"}), 404
    return send_file(
        str(path), as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/reports")
def api_reports():
    files = sorted(
        REPORTS_DIR.glob("*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    return jsonify([
        {
            "filename": f.name,
            "size_kb":  round(f.stat().st_size / 1024, 1),
            "modified": datetime.fromtimestamp(f.stat().st_mtime).strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        }
        for f in files
    ])


@app.route("/api/folders")
def api_folders():
    s, t = get_dirs()
    return jsonify({
        "source_dir":  str(s),
        "target_dir":  str(t),
        "reports_dir": str(REPORTS_DIR),
    })


@app.route("/api/clear-results", methods=["POST"])
def api_clear_results():
    results_store.clear()
    file_states.clear()
    activity_log.clear()
    scan_status["last_scan"]       = None
    scan_status["scanning"]        = False
    scan_status["error"]           = None
    scan_status["current_file"]    = None
    scan_status["total_files"]     = 0
    scan_status["completed_files"] = 0
    log_event("Results cleared manually", "info")
    return jsonify({"ok": True})


if __name__ == "__main__":
    s, t = get_dirs()
    cfg  = load_config()

    print("\n  Genpact SAP Validator - Dashboard")
    print(f"  Source dir     -> {s}")
    print(f"  Target dir     -> {t}")
    print(f"  Reports        -> {REPORTS_DIR}")
    print(f"  Pass threshold -> {cfg.get('pass_threshold', 100)}%")
    print("  Open           -> http://localhost:5000\n")

    # Print all known SAP field labels on startup so you can verify
    print(f"  SAP field dictionary: {len(SAP_FIELD_LABELS)} known fields")
    print(f"  e.g. KUNNR={SAP_FIELD_LABELS.get('KUNNR')}  "
          f"MATNR={SAP_FIELD_LABELS.get('MATNR')}  "
          f"KTOKD={SAP_FIELD_LABELS.get('KTOKD')}\n")

    threading.Thread(target=scan_and_validate_all, daemon=True).start()

    # threading.Thread(target=background_watcher, args=(60,), daemon=True).start()

    app.run(debug=False, port=5000, use_reloader=False)
